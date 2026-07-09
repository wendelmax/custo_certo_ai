import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

def exportar_analise_excel(caminho_saida: str, df_margens: pd.DataFrame, df_absorcao_vol: pd.DataFrame, 
                           df_absorcao_horas: pd.DataFrame, df_absorcao_custo: pd.DataFrame, 
                           df_ineficiencias: pd.DataFrame, pe_dict: dict, custos_fixos: float) -> None:
    wb = Workbook()
    
    # Estilos visuais
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Azul Escuro
    zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid") # Cinza Claro
    bold_font = Font(name="Calibri", size=11, bold=True)
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # 1. ABA 1: Dashboard C-Level
    ws_dash = wb.active
    ws_dash.title = "Dashboard C-Level"
    ws_dash.sheet_view.showGridLines = True
    
    ws_dash.append(["Métrica Executiva", "Valor Financeiro"])
    ws_dash.append(["Faturamento Total da Fábrica", pe_dict.get("faturamento_atual_total", 0.0)])
    ws_dash.append(["Custos Fixos Mensais Fabris", custos_fixos])
    ws_dash.append(["Ponto de Equilíbrio Geral", pe_dict.get("faturamento_equilibrio_geral", 0.0)])
    ws_dash.append(["Margem de Contribuição Média (%)", pe_dict.get("margem_contribuição_media_ponderada", 0.0)])
    ws_dash.append(["Custo Total de Ineficiências (Refugo + Paradas)", df_ineficiencias["perda_ineficiencia_total"].sum() if "perda_ineficiencia_total" in df_ineficiencias.columns else 0.0])
    
    # Estilizar Aba 1
    for col in ws_dash.iter_cols(min_row=1, max_row=6, min_col=1, max_col=2):
        for cell in col:
            cell.border = thin_border
    for cell in ws_dash[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row in range(2, 7):
        ws_dash.cell(row=row, column=2).number_format = 'R$ #,##0.00'
        if row == 5:
            ws_dash.cell(row=row, column=2).number_format = '0.0%'
    ws_dash.cell(row=6, column=2).font = Font(name="Calibri", size=11, bold=True, color="FF0000") # Custo ineficiencia em vermelho
    
    # Helper para exportar tabelas genéricas
    def exportar_planilha_tabela(nome_aba: str, df: pd.DataFrame, formats: dict):
        ws = wb.create_sheet(title=nome_aba)
        ws.sheet_view.showGridLines = True
        
        # Cabeçalhos
        headers = list(df.columns)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
        # Conteúdo
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
            ws.append(row)
            # Estilo zebra e bordas
            fill = zebra_fill if r_idx % 2 == 0 else PatternFill(fill_type=None)
            for c_idx in range(1, len(row) + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.border = thin_border
                if fill.fill_type:
                    cell.fill = fill
                    
                # Formatação específica de colunas
                col_name = headers[c_idx - 1]
                if col_name in formats:
                    cell.number_format = formats[col_name]
                    
    # Formatos das colunas
    formats_margem = {
        "preco_venda_unitario": 'R$ #,##0.00',
        "custo_variavel_unitario": 'R$ #,##0.00',
        "despesas_variaveis_unit": 'R$ #,##0.00',
        "margem_contribuição_unitaria": 'R$ #,##0.00',
        "margem_contribuição_percentual": '0.0%',
        "margem_contribuição_total": 'R$ #,##0.00',
        "volume_vendas_mensal": '#,##0'
    }
    formats_absorcao = {
        "preco_venda_unitario": 'R$ #,##0.00',
        "custo_variavel_unitario": 'R$ #,##0.00',
        "despesas_variaveis_unit": 'R$ #,##0.00',
        "quantidade_produzida": '#,##0',
        "horas_maquina_ativas": '#,##0.0',
        "rateio_custo_fixo": 'R$ #,##0.00',
        "custo_fixo_unitario": 'R$ #,##0.00',
        "custo_absorcao_unitario": 'R$ #,##0.00',
        "lucro_unitario_absorcao": 'R$ #,##0.00',
        "lucro_total_absorcao": 'R$ #,##0.00'
    }
    formats_ineficiencias = {
        "quantidade_produzida": '#,##0',
        "quantidade_refugada": '#,##0',
        "taxa_refugo_percentual": '0.0%',
        "horas_maquina_parada": '#,##0.0',
        "custo_refugo": 'R$ #,##0.00',
        "custo_parada_maquina": 'R$ #,##0.00',
        "perda_ineficiencia_total": 'R$ #,##0.00'
    }
    
    # 2. ABA 2: Margens
    exportar_planilha_tabela("Margens e PE", df_margens, formats_margem)
    
    # 3. ABA 3: Absorção
    # Concatenar visões de absorção para relatório side-by-side limpo
    abs_consolidado = df_absorcao_vol[["produto_id", "nome_produto", "preco_venda_unitario", "custo_absorcao_unitario", "lucro_unitario_absorcao"]].copy()
    abs_consolidado.columns = ["produto_id", "nome_produto", "Preço Venda", "Absorção Unit. (Volume)", "Lucro Unit. (Volume)"]
    
    df_horas_sub = df_absorcao_horas[["produto_id", "custo_absorcao_unitario", "lucro_unitario_absorcao"]].copy()
    df_horas_sub.columns = ["produto_id", "Absorção Unit. (Horas)", "Lucro Unit. (Horas)"]
    
    df_custo_sub = df_absorcao_custo[["produto_id", "custo_absorcao_unitario", "lucro_unitario_absorcao"]].copy()
    df_custo_sub.columns = ["produto_id", "Absorção Unit. (Custo Dir)", "Lucro Unit. (Custo Dir)"]
    
    abs_consolidado = pd.merge(abs_consolidado, df_horas_sub, on="produto_id", how="left")
    abs_consolidado = pd.merge(abs_consolidado, df_custo_sub, on="produto_id", how="left")
    
    formats_consolidado = {
        "Preço Venda": 'R$ #,##0.00',
        "Absorção Unit. (Volume)": 'R$ #,##0.00',
        "Lucro Unit. (Volume)": 'R$ #,##0.00',
        "Absorção Unit. (Horas)": 'R$ #,##0.00',
        "Lucro Unit. (Horas)": 'R$ #,##0.00',
        "Absorção Unit. (Custo Dir)": 'R$ #,##0.00',
        "Lucro Unit. (Custo Dir)": 'R$ #,##0.00'
    }
    exportar_planilha_tabela("Rateios por Absorção", abs_consolidado, formats_consolidado)
    
    # 4. ABA 4: Ineficiências
    exportar_planilha_tabela("Ineficiências Fabris", df_ineficiencias, formats_ineficiencias)
    
    # Auto-fit columns nas abas
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 8, 15)
            
    wb.save(caminho_saida)
